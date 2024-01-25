import openpyxl
from openpyxl.worksheet import worksheet
from openpyxl.styles import Alignment, Border, Side, Font, colors
import calendar
import shutil
import sys

_FILE = "calendar.xlsx"
_SHEET_NAME = "Sheet1"
_YEAR_TO_OFFSET = 2

_MONTH_NAMES_ITA = [
    "Gennaio",
    "Febbraio",
    "Marzo",
    "Aprile",
    "Maggio",
    "Giugno",
    "Luglio",
    "Agosto",
    "Settembre",
    "Ottobre",
    "Novembre",
    "Dicembre",
]

_DAY_NAMES_ITA = [
    "Lunedì",
    "Martedì",
    "Mercoledì",
    "Giovedì",
    "Venerdì",
    "Sabato",
    "Domenica",
]

_BLOCK_ROW_SIZE = 33
_BLOCK_COL_SIZE = 7

_HEADER_ROW_SIZE = 6
_HEADER_COL_SIZE = 6

_NUMBER_BOX_ROW_SIZE = 6
_NUMBER_BOX_COL_SIZE = 2

_MONTH_ROW_OFFSET = 1
_MONTH_ROW_SIZE = _MONTH_ROW_OFFSET + 1
_MONTH_COL_OFFSET = 3
_MONTH_COL_SIZE = _MONTH_COL_OFFSET + 3

_DAY_ROW_OFFSET = 4
_DAY_ROW_SIZE = _DAY_ROW_OFFSET
_DAY_COL_OFFSET = 4
_DAY_COL_SIZE = _DAY_COL_OFFSET + 1


class CellRange:
    def __init__(self, start_row: int, start_col: int, end_row: int, end_col: int):
        self.start_row = start_row
        self.start_col = start_col
        self.end_row = end_row
        self.end_col = end_col


class Header:
    def __init__(self, day_number: int, day_name: str, month: str):
        self.day_number = day_number
        self.day_name = day_name
        self.month = month


def create_block(
    sheet: worksheet.Worksheet,
    day: int,
    day_name: str,
    month: str,
    start_row: int,
    start_col: int,
):
    header = Header(day, day_name, month)
    sheet = create_header(sheet, header, start_row, start_col)
    sheet = create_body(sheet, start_row + 7, start_col)


def merge_and_center(
    sheet: worksheet.Worksheet,
    cellRange: CellRange,
    value: any = None,
    font: Font = None,
) -> worksheet.Worksheet:
    if value is not None:
        sheet.cell(cellRange.start_row, cellRange.start_col, value)
    sheet.merge_cells(
        None,
        cellRange.start_row,
        cellRange.start_col,
        cellRange.end_row,
        cellRange.end_col,
    )
    cell = sheet.cell(cellRange.start_row, cellRange.start_col)
    if font is not None:
        cell.font = font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell

    return sheet


def create_body(
    sheet: worksheet.Worksheet, start_row: int, start_col: int
) -> worksheet.Worksheet:
    sheet = create_box_border(
        sheet, start_row, start_col, start_row + 25, start_col + 6
    )
    return sheet


def create_header(
    sheet: worksheet.Worksheet, values: Header, start_row: int, start_col: int
) -> worksheet.Worksheet:
    header = CellRange(
        start_row, start_col, start_row + _HEADER_ROW_SIZE, start_col + _HEADER_COL_SIZE
    )
    numberBox = CellRange(
        start_row,
        start_col,
        start_row + _NUMBER_BOX_ROW_SIZE,
        start_col + _NUMBER_BOX_COL_SIZE,
    )
    monthBox = CellRange(
        start_row + _MONTH_ROW_OFFSET,
        start_col + _MONTH_COL_OFFSET,
        start_row + _MONTH_ROW_SIZE,
        start_col + _MONTH_COL_SIZE,
    )
    dayBox = CellRange(
        start_row + _DAY_ROW_OFFSET,
        start_col + _DAY_COL_OFFSET,
        start_row + _DAY_ROW_SIZE,
        start_col + _DAY_COL_SIZE,
    )

    # number box
    font = Font(size=72, color=colors.Color(rgb=colors.BLACK))
    sheet = merge_and_center(sheet, numberBox, values.day_number, font)
    cell = sheet.cell(numberBox.start_row, numberBox.start_col)
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # month box
    font = Font(size=28, color=colors.Color(rgb=colors.BLACK))
    sheet = merge_and_center(sheet, monthBox, values.month, font)

    # day name box
    bold = False
    if values.day_name == "Sabato" or values.day_name == "Domenica":
        bold = True
    font = Font(size=14, color=colors.Color(rgb=colors.BLACK), bold=bold)
    sheet = merge_and_center(sheet, dayBox, values.day_name, font)

    sheet = create_box_border(
        sheet, header.start_row, header.start_col, header.end_row, header.end_col
    )
    return sheet


def create_box_border(
    sheet: worksheet.Worksheet,
    start_row: int,
    start_col: int,
    end_row: int,
    end_col: int,
) -> worksheet.Worksheet:
    # Crea un bordo esterno attorno al range di celle
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)

            # Seleziona solo i bordi esterni del range
            border = Border()
            if row == start_row:
                border.top = Side(style="thin")
            if row == end_row:
                border.bottom = Side(style="thin")
            if col == start_col:
                border.left = Side(style="thin")
            if col == end_col:
                border.right = Side(style="thin")

            cell.border = border
    return sheet


def create_new_calendar_file(file_path_origin: str, file_path_destination: str):
    try:
        shutil.copy2(file_path_origin, file_path_destination)
        print(f"File copied from {file_path_origin} to {file_path_destination}")
    except FileNotFoundError:
        print(f"Error: File {file_path_origin} not exist.")
    except PermissionError:
        print(f"Error: Permission error.")


def get_year_from_command_line() -> int:
    if len(sys.argv) > 1:
        return int(sys.argv[1])
    else:
        print("Please insert year.")
        exit()


def main():
    start_row = 1
    start_col = 1
    month_limit = 13

    try:
        year_from = get_year_from_command_line()
    except ValueError:
        print("Invalid year format.")
        exit()

    file_name = f"calendar_{year_from}.xlsx"
    create_new_calendar_file(_FILE, file_name)

    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[_SHEET_NAME]

    year_to = year_from + _YEAR_TO_OFFSET
    side = True

    for year in range(year_from, year_to):
        # adjust the month limit for the last year
        if year == year_to - 1:
            month_limit = 4  # stop on March (3)

        for month in range(1, month_limit):
            month_name = _MONTH_NAMES_ITA[month - 1]
            weeks_day = calendar.monthcalendar(year, month)

            for week in weeks_day:
                for day in week:
                    if day != 0:
                        num_day_week = calendar.weekday(year, month, day)
                        day_name = _DAY_NAMES_ITA[num_day_week]

                        # draw day block in the sheet
                        create_block(
                            sheet, day, day_name, month_name, start_row, start_col
                        )

                        # next block position
                        if side:
                            start_col += _BLOCK_COL_SIZE
                        else:
                            start_row += _BLOCK_ROW_SIZE
                            start_col = 1
                        side = not side

    workbook.save(file_name)


if __name__ == "__main__":
    main()
